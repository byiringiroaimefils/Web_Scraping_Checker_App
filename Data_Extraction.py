import requests
from bs4 import BeautifulSoup
import openpyxl

url = "https://www.newtimes.co.rw"
response = requests.get(url)


soup = BeautifulSoup(response.content, 'html.parser')

title_articles= soup.find_all("h1", class_="title article")
article_bodies= soup.find_all("div", class_="article-body")

workbook = openpyxl.Workbook()
sheet = workbook.active


sheet['A1'] = 'Title Article'
sheet['B1'] = 'article Body'

for i, p in  enumerate(zip(title_articles, article_bodies), start=2) :
        sheet.cell(row=i+1, column=1, value=p.text.strip())
        sheet.cell(row=i+1, column=2, value=p.text.strip())

workbook.save('TheNew_data.xlsx')
print("Data has been extracted ")

