import requests
from bs4 import BeautifulSoup
import openpyxl

url = "https://igihe.com/index.php"
response = requests.get(url)


soup = BeautifulSoup(response.content, 'html.parser')

paragraphs = soup.find_all('p')

workbook = openpyxl.Workbook()
sheet = workbook.active


sheet['A1'] = 'Paragraph Number'
sheet['B1'] = 'Text'

for i, p in enumerate(paragraphs):
    sheet.cell(row=i+1, column=1, value=i)
    sheet.cell(row=i+1, column=2, value=p.text)

workbook.save('igihe_data.xlsx')
print("Data has been extracted ")

